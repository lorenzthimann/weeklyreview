# working w openpyxl this time

import openpyxl
# WEEK NB IS SET TO += 15 !!!
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from datetime import date

alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']

wb = openpyxl.load_workbook('weekly_review.xlsx')


# grab the active worksheet
ws = wb.active
# creating the week in the lines

i_for_content = 35
i_for_cell = 2
while i_for_content < 53:
    i_for_cell = str(i_for_cell)
    cell = "A" + i_for_cell
    i_for_cell = int(i_for_cell)
    ws[cell] = i_for_content
    i_for_cell += 1
    i_for_content += 1




# put the habits on the 1st line
habits = ["Carnegie", "Drink", "Use transportation", "S podcasts & video", "Intresting TgB", "Beautiful notebooks", "fr./g.", "c*n*", "wllm", "passionate + part.", "Music"]
index_for_habits = 0
index_for_alphabet = 0
while index_for_habits < len(habits):
    cell = str(alphabet[index_for_alphabet + 1].upper()) + "1"
    ws[cell] = habits[index_for_habits]
    index_for_alphabet += 1
    index_for_habits += 1

ws["B1"] = habits[0]
ws["C1"] = habits[1]
ws["D1"] = habits[2]
ws["E1"] = habits[3]
ws["F1"] = habits[4]
ws["G1"] = habits[5]
ws["H1"] = habits[6]
ws["I1"] = habits[7]
ws["J1"] = habits[8]
ws["K1"] = habits[9]
ws["L1"] = habits[10]




week_number = weekNumber = date.today().isocalendar()[1] + 15 # to change to 0
carnegie = input(f"{habits[0]}  : ")
drink = input(f"{habits[1]}  : ")
use_transportation = input(f"{habits[2]}  : ")
s_podcast_and_video = input(f"{habits[3]} : ")
tgb = input(f"{habits[4]}  : ")
notebooks = input(f"{habits[5]}  : ")
fr_g = input(f"{habits[6]}  : ")
c_n = input(f"{habits[7]}  : ")
wllm = input(f"{habits[8]}  : ")
passionate = input(f"{habits[9]}  : ")
music = input(f"{habits[10]}  : ")
todays_line = int(week_number) - 33 # to change as well
evaluations = [carnegie, drink, use_transportation, s_podcast_and_video,  tgb, notebooks,  fr_g, c_n,  wllm, passionate, music]

index = 0
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
orangeFill = PatternFill(start_color='F79D00',
                   end_color='F79D00',
                   fill_type='solid')
blueFill = PatternFill(start_color='0A9EFF',
                   end_color='0A9EFF',
                   fill_type='solid')
greenFill = PatternFill(start_color='98F50E',
                   end_color='98F50E',
                   fill_type='solid')

while index < len(evaluations):
    evaluation = evaluations[index]
    cell_to_handle = alphabet[index + 1].upper() + str(todays_line)
    if evaluation == "1":
       ws[cell_to_handle].fill = redFill
    elif evaluation == "2":
        ws[cell_to_handle].fill = orangeFill
    elif evaluation == "3":
        ws[cell_to_handle].fill = blueFill
    elif evaluation == "4":
        ws[cell_to_handle].fill = greenFill
    index += 1
wb.save("weekly_review.xlsx")
# create new ws for txts habits

